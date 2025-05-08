#!/usr/bin/env python3
# Merges Excel files from RP Data into a single file

import os
import logging
import pandas as pd
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from collections import Counter
import time

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(name)s - %(message)s'
)
logger = logging.getLogger(__name__)

def check_cancellation(progress_callback, current_progress, message):
    """Helper function to check if job has been cancelled"""
    if progress_callback is None:
        return False
        
    # Call the progress callback and check its return value
    return progress_callback(current_progress, message) is False

def find_header_row(df):
    """Find the row index where 'Property Photo' appears in the first column."""
    for idx, value in enumerate(df.iloc[:, 0]):
        if str(value).strip() == 'Property Photo':
            return idx
    return None

def get_hyperlink_from_excel(file_path, sheet_name=0, row_idx=None, column_name="Open in RPData"):
    """
    Directly extract hyperlinks from Excel cells using openpyxl.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (int/str): Sheet index or name
        row_idx (int): Row index in the dataframe (0-based)
        column_name (str): Name of the column containing hyperlinks
        
    Returns:
        dict: Mapping of row index to hyperlink URL
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Get the sheet
        if isinstance(sheet_name, int):
            sheet = wb.worksheets[sheet_name]
        else:
            sheet = wb[sheet_name]
        
        # Find header row for Property Photo
        header_row = None
        for row_num, row in enumerate(sheet.iter_rows(), 1):
            for cell in row:
                if cell.value == "Property Photo":
                    header_row = row_num
                    break
            if header_row:
                break
        
        if not header_row:
            logger.error(f"Could not find 'Property Photo' header in {file_path}")
            return {}
        
        # Find column index for 'Open in RPData'
        col_idx = None
        for col_num, cell in enumerate(sheet[header_row], 1):
            if cell.value == column_name:
                col_idx = col_num
                break
        
        if not col_idx:
            logger.warning(f"Could not find '{column_name}' column in {file_path}")
            return {}
        
        # Get hyperlinks from cells
        hyperlinks = {}
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=header_row+1), header_row+1):
            cell = sheet.cell(row=row_num, column=col_idx)
            
            # Extract URL from cell
            url = None
            
            # Check for hyperlink attribute
            if cell.hyperlink:
                url = cell.hyperlink.target
            
            # Check for formula
            elif cell.value and isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK('):
                match = re.search(r'=HYPERLINK\("([^"]+)"', cell.value)
                if match:
                    url = match.group(1)
            
            # Store URL if found
            if url:
                df_row_idx = row_num - (header_row + 1)
                hyperlinks[df_row_idx] = url
        
        return hyperlinks
    
    except Exception as e:
        logger.error(f"Error extracting hyperlinks from {file_path}: {e}")
        return {}

def extract_hyperlink(cell_content):
    """Extract hyperlink URL from Excel HYPERLINK formula or direct URL."""
    # Handle NaN values
    if pd.isna(cell_content):
        return ""
        
    # Handle Excel formula format
    if isinstance(cell_content, str) and cell_content.startswith('=HYPERLINK('):
        match = re.search(r'=HYPERLINK\("([^"]+)"', cell_content)
        if match:
            return match.group(1)
    
    # If it's already a URL string
    if isinstance(cell_content, str) and (cell_content.startswith('http://') or cell_content.startswith('https://')):
        return cell_content
            
    return cell_content

def generate_filename(locations, property_types, min_floor, max_floor, output_dir=None):
    """
    Generate a filename based on search criteria and current date/time.
    
    Args:
        locations (list): List of locations searched
        property_types (list): List of property types searched
        min_floor (str): Minimum floor size
        max_floor (str): Maximum floor size
        output_dir (str): Optional output directory
        
    Returns:
        str: Generated filename
    """
    # Format current date and time
    now = datetime.now()
    date_time_str = now.strftime("%d_%m_%Y_%H_%M")
    
    # Format locations for filename (use first word of each location)
    location_str = "_".join([loc.split()[0].replace(",", "") for loc in locations[:2]])
    
    # Create output directory if it doesn't exist and not provided
    if output_dir is None:
        output_dir = "merged_properties"
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Create filename with output directory
    filename = f"{output_dir}/Properties_{location_str}_{date_time_str}.xlsx"
    
    return filename

def process_excel_files(files_dict, locations, property_types, min_floor, max_floor, business_type, 
                        headless=False, output_dir=None, output_file=None, progress_callback=None):
    """
    Process and merge RP Data Excel files.
    
    Args:
        files_dict (dict): Dictionary with search types as keys and file paths as values
        locations (list): List of locations searched
        property_types (list): List of property types searched
        min_floor (str): Minimum floor size
        max_floor (str): Maximum floor size
        business_type (str): Type of business searched either Vet or Health
        headless (bool): Whether to run in headless mode
        output_dir (str): Job-specific output directory for merged file
        output_file (str, optional): Specific output file path. If None, will be generated.
        progress_callback (function): Callback function for progress updates
        
    Returns:
        bool: Success status
    """
    # Default progress reporting function if none provided
    if progress_callback is None:
        def progress_callback(percentage, message):
            pass  # No-op if no callback provided
            return True  # Always continue
    
    logger.info("===== PROCESSING EXCEL FILES =====")

    # Check for cancellation at the start
    if check_cancellation(progress_callback, 96, "Starting to process Excel files..."):
        logger.info("Job cancelled at the start of Excel processing")
        return False
    
    # Generate filename if not provided
    if output_file is None:
        output_file = generate_filename(locations, property_types, min_floor, max_floor, output_dir)
    
    try:
        # Check for cancellation before collecting data
        if check_cancellation(progress_callback, 96, "Parsing Excel files..."):
            logger.info("Job cancelled before Excel parsing")
            return False
        
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        
        # Set worksheet name to "Properties" as requested
        ws.title = "Properties"
        
        # Create headers for the merged file
        headers = [
            "Type", "Property Photo", "Street Address", "Suburb", "State", "Postcode",
            "Site Zoning", "Property Type", "Bed", "Bath", "Car", "Extra Cost for Parks", "Land Size (m²)",
            "Floor Size (m²)", "Year Built", "Agency", "Agent", "Contact Phone", "Email",
            "Contacted (T/F)", "Land Use", "Development Zone", "Parcel Details", "Owner Type",
            "Website Link",
            # For "Sales" only
            "Sale Price", "Sale Date", "Settlement Date", "Sale Type", "Owner 1 Name",
            "Owner 2 Name", "Owner 3 Name", "Vendor 1 Name", "Vendor 2 Name", "Vendor 3 Name",
            # For "For Sale" only
            "First Listed Price", "First Listed Date", "Last Listed Price", "Last Listed Date",
            "Listing Type",
            # For "For Rent" only
            "First Rental Price", "First Rental Date", "Last Rental Price", "Last Rental Date",
            "Outgoings Ex GST", "Total Lease Price (Base + Outgoings)",
            # For "For Rent" and "For Sale" only
            "Days on Market", "Active Listing",
            # Additional blank columns
            "Comments Y=Recommended, E=Evaluating, R=Rejected", "Date Added", "Date Presented",
            "Allowable Use in Zone (T/F)", "$/m²", "Available (T/F)", "Suitable (T/F)",
            "PUT IN REPORT (T/F)", "Client Feedback", "Busi's Comment"
        ]
        
        # Add headers to the first row
        ws.append(headers)
        
        # Make the header row bold
        for cell in ws[1]:
            cell.font = Font(bold=True, size=14)

        # First pass: collect all rows
        all_rows = []
        
        logger.info("First pass: collecting all property data...")
        
        progress_callback(97, "Extracting data from forSale, forRent and Sales...")

        row_index = 0
        for search_type, file_path in files_dict.items():
            logger.info(f"Reading file for {search_type}: {file_path}")
            
            try:
                # First get hyperlinks directly from Excel file
                hyperlinks = get_hyperlink_from_excel(file_path)
                logger.info(f"Retrieved {len(hyperlinks)} hyperlinks directly from {file_path}")
                
                # Read the Excel file
                df = pd.read_excel(file_path, engine='openpyxl')
                
                # Find the header row
                header_row = find_header_row(df)
                
                if header_row is None:
                    logger.error(f"Could not find header row in {file_path}")
                    continue
                
                # Set the header row as the column names and get data
                df.columns = df.iloc[header_row]
                data_df = df.iloc[header_row + 1:].reset_index(drop=True)
                
                # Log column headers for debugging
                logger.info(f"Columns in {search_type} file: {list(data_df.columns)}")
                
                # Process each row
                for idx, row in data_df.iterrows():
                    new_row = ["N/A"] * len(headers)  # Initialize with N/A for all columns
                    
                    # Set type based on search type
                    if search_type == "Sales":
                        new_row[0] = "Sold"
                    elif search_type == "For Sale":
                        new_row[0] = "For Sale"
                    elif search_type == "For Rent":
                        # Check Active Listing column for For Rent
                        active_listing = row.get("Active Listing")
                        new_row[0] = "For Lease" if str(active_listing) == "True" else "Already Leased"
                    
                    # Common fields across all file types
                    new_row[1] = ""  # Property Photo - Now leaving blank
                    
                    # Extract address components
                    street_address = str(row.get("Street Address", ""))
                    suburb = str(row.get("Suburb", ""))
                    state = str(row.get("State", ""))
                    postcode = str(row.get("Postcode", ""))
                    
                    new_row[2] = street_address
                    new_row[3] = suburb
                    new_row[4] = state
                    new_row[5] = postcode
                    
                    # Just set Site Zoning to empty string
                    new_row[6] = ""
                    
                    new_row[7] = row.get("Property Type", "")
                    new_row[8] = row.get("Bed", "")
                    new_row[9] = row.get("Bath", "")
                    new_row[10] = row.get("Car", "")
                    
                    # Fields that should be blank
                    new_row[11] = ""  # Extra Cost for Parks
                    
                    # Land Size and Floor Size with correct symbol
                    new_row[12] = row.get("Land Size (m²)", "")
                    new_row[13] = row.get("Floor Size (m²)", "")
                    new_row[14] = row.get("Year Built", "")
                    new_row[15] = row.get("Agency", "")
                    new_row[16] = row.get("Agent", "")
                    
                    # Fields that should be blank - contact details are not scraped now
                    new_row[17] = ""  # Contact Phone
                    new_row[18] = ""  # Email
                    new_row[19] = ""  # Contacted (T/F)
                    
                    # Common fields across all file types
                    new_row[20] = row.get("Land Use", "")
                    new_row[21] = row.get("Development Zone", "")
                    new_row[22] = row.get("Parcel Details", "")
                    new_row[23] = row.get("Owner Type", "")
                    
                    # Website Link from Open in RPData - multi-step approach to get the link
                    # First try using the hyperlinks extracted directly from Excel
                    if idx in hyperlinks:
                        # We found a direct hyperlink from the Excel file
                        new_row[24] = hyperlinks[idx]
                        logger.info(f"Found hyperlink directly from Excel: {hyperlinks[idx]}")
                    else:
                        # Otherwise try different approaches for getting the link
                        rp_data_link = None
                        
                        # Method 1: Try direct column access
                        if "Open in RPData" in row:
                            rp_data_link = row["Open in RPData"]
                        
                        # Method 2: Look for column with RPData or Link in its name
                        if pd.isna(rp_data_link) or not rp_data_link:
                            for col in row.index:
                                if "RPData" in col or "Link" in col:
                                    if not pd.isna(row[col]) and row[col]:
                                        rp_data_link = row[col]
                                        break
                        
                        # Method 3: If it's still NaN but we have a street address, create a generic URL
                        if (pd.isna(rp_data_link) or not rp_data_link) and street_address:
                            # Clean the address for a URL-friendly format
                            clean_addr = f"{street_address.lower().replace(' ', '-')}-{suburb.lower().replace(' ', '-')}-{state.lower()}-{postcode}"
                            rp_data_link = f"https://rpp.corelogic.com.au/property/{clean_addr}"
                            logger.warning(f"Created generic URL for {street_address}: {rp_data_link}")
                        
                        # Extract hyperlink and set it
                        new_row[24] = extract_hyperlink(rp_data_link) if not pd.isna(rp_data_link) else ""
                        
                        # As a last resort, preserve the raw formula
                        if not new_row[24] and isinstance(rp_data_link, str) and "HYPERLINK" in rp_data_link:
                            # Just use the raw formula - we'll extract the URL during export
                            new_row[24] = rp_data_link
                    
                    # Fields specific to "Sales"
                    if search_type == "Sales":
                        new_row[25] = row.get("Sale Price", "")
                        new_row[26] = row.get("Sale Date", "")
                        new_row[27] = row.get("Settlement Date", "")
                        new_row[28] = row.get("Sale Type", "")
                        new_row[29] = row.get("Owner 1 Name", "")
                        new_row[30] = row.get("Owner 2 Name", "")
                        new_row[31] = row.get("Owner 3 Name", "")
                        new_row[32] = row.get("Vendor 1 Name", "")
                        new_row[33] = row.get("Vendor 2 Name", "")
                        new_row[34] = row.get("Vendor 3 Name", "")
                    
                    # Fields specific to "For Sale"
                    if search_type == "For Sale":
                        new_row[35] = row.get("First Listed Price", "")
                        new_row[36] = row.get("First Listed Date", "")
                        new_row[37] = row.get("Last Listed Price", "")
                        new_row[38] = row.get("Last Listed Date", "")
                        new_row[39] = row.get("Listing Type", "")
                        
                        # For Sale also has Days on Market and Active Listing
                        new_row[46] = row.get("Days on Market", "")
                        new_row[47] = row.get("Active Listing", "")
                    
                    # Fields specific to "For Rent"
                    if search_type == "For Rent":
                        new_row[40] = row.get("First Rental Price", "")
                        new_row[41] = row.get("First Rental Date", "")
                        new_row[42] = row.get("Last Rental Price", "")
                        new_row[43] = row.get("Last Rental Date", "")
                        new_row[44] = row.get("Outgoings Ex GST", "")
                        new_row[45] = row.get("Total Lease Price (Base + Outgoings)", "")
                        
                        # For Rent also has Days on Market and Active Listing
                        new_row[46] = row.get("Days on Market", "")
                        new_row[47] = row.get("Active Listing", "")
                    
                    # Make additional columns blank instead of N/A
                    for i in range(48, len(headers)):
                        new_row[i] = ""
                    
                    # Add the new row to our collection
                    all_rows.append(new_row)
                    row_index += 1
                
                logger.info(f"Collected {len(data_df)} rows from {search_type}")
            
            except Exception as e:
                logger.error(f"Error processing file {file_path}: {e}")
                import traceback
                logger.error(traceback.format_exc())

        if check_cancellation(progress_callback, 97, "Processing property data..."):
            logger.info("Job cancelled while processing property data")
            return False
        
        progress_callback(98, "Writing all properties to final merged file...")

        # Write all rows to the worksheet
        logger.info("Writing all rows to the output file...")
        
        # Current date for "Date Added" column
        today_str = datetime.now().strftime("%d/%m/%Y")

        if check_cancellation(progress_callback, 98, "Writing to Excel file..."):
            logger.info("Job cancelled before writing Excel file")
            return False
        
        for i, row_data in enumerate(all_rows, 2):  # Start from row 2 (after headers)
            # Set "Property Photo" column to empty (column B, index 1)
            row_data[1] = ""  # Clear any existing image URL
            
            # Set "Date Added" column to today's date (column AW, index 49)
            row_data[49] = today_str
            
            # Write all values for this row
            for j, value in enumerate(row_data):
                # For Website Link column (25th column, index 24)
                if j == 24 and isinstance(value, str) and (value.startswith('http://') or value.startswith('https://')):
                    cell = ws.cell(row=i, column=j+1, value=value)
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
                else:
                    # Regular value
                    ws.cell(row=i, column=j+1, value=value)
        
        # Set font size for all cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row == 1:
                    # Header row - bold and size 14
                    cell.font = Font(bold=True, size=14)
                else:
                    # Data rows - size 12
                    cell.font = Font(size=12)
        
        # Adjust column widths
        col_widths = {
            'A': 15,  # Type
            'B': 15,  # Property Photo (empty now)
            'C': 30,  # Street Address
            'D': 20,  # Suburb
            'E': 10,  # State
            'F': 10,  # Postcode
            'G': 35,  # Site Zoning
            'H': 20,  # Property Type
            'I': 8,   # Bed
            'J': 8,   # Bath
            'K': 8,   # Car
            'L': 20,  # Extra Cost for Parks
            'M': 15,  # Land Size
            'N': 15,  # Floor Size
            'O': 12,  # Year Built
            'P': 20,  # Agency
            'Q': 20,  # Agent
            'R': 15,  # Contact Phone
            'S': 25,  # Email
            'T': 15,  # Contacted
            'U': 20,  # Land Use
            'V': 20,  # Development Zone
            'W': 20,  # Parcel Details
            'X': 15,  # Owner Type
            'Y': 30,  # Website Link
            'AZ': 10, # Allowable Use in Zone
        }
        
        # Apply column widths
        for col_letter, width in col_widths.items():
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = width
        
        # Set default width for other columns
        for col in ws.columns:
            col_letter = col[0].column_letter
            if col_letter not in col_widths:
                ws.column_dimensions[col_letter].width = 15
        
        # Save the merged file
        wb.save(output_file)
        logger.info(f"Merged file saved to {output_file}")
        return True
    
    except Exception as e:
        logger.error(f"Error processing Excel files: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

# Test function to allow running the module directly for testing
def test_merge_excel():
    """Test function for the merge_excel module with job-specific directories."""
    print("Testing merge_excel module...")
    
    # Create a test job ID
    test_job_id = f"test_{int(time.time())}"
    
    # Create job-specific directories
    job_download_dir = os.path.join("downloads", test_job_id)
    job_output_dir = os.path.join("merged_properties", test_job_id)
    
    os.makedirs(job_download_dir, exist_ok=True)
    os.makedirs(job_output_dir, exist_ok=True)
    
    # Check if test files exist in the job download directory
    test_files = {
        "Sales": os.path.join(job_download_dir, "recentSaleExport_test.xlsx"),
        "For Sale": os.path.join(job_download_dir, "forSaleExport_test.xlsx"),
        "For Rent": os.path.join(job_download_dir, "forRentExport_test.xlsx")
    }
    
    # Check if at least one test file exists
    file_exists = False
    for file_type, file_path in test_files.items():
        if os.path.exists(file_path):
            file_exists = True
            break
    
    if not file_exists:
        print("No test files found. Please create test files first.")
        return False
    
    # Search criteria
    locations = ["Test Location 1", "Test Location 2"]
    property_types = ["Business", "Commercial"]
    min_floor = "Min"
    max_floor = "1200"
    business_type = "Vet"
    
    # Run the process with job-specific output directory
    result = process_excel_files(
        files_dict=test_files,
        locations=locations,
        property_types=property_types,
        min_floor=min_floor,
        max_floor=max_floor,
        business_type=business_type,
        headless=True,
        output_dir=job_output_dir
    )
    
    return result

if __name__ == "__main__":
    # Run the test function when this script is executed directly
    test_merge_excel()