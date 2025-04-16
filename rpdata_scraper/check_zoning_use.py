#!/usr/bin/env python3
# Checks if a property's zoning allows specific business uses (Vet or Health)
# Uses the "Allowable Use in the Zone - TABLE.xlsx" reference file

import os
import logging
from openpyxl import load_workbook

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_absolute_file_path(relative_filename):
    """
    Resolve the absolute path of a file, searching in multiple potential locations.
    
    Args:
        relative_filename (str): Filename to locate
    
    Returns:
        str: Absolute path to the file, or None if not found
    """
    # Potential search paths
    search_paths = [
        # Current working directory
        os.getcwd(),
        
        # Directory of the current script
        os.path.dirname(os.path.abspath(__file__)),
        
        # Parent directory of current script
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        
        # Specific for your project structure
        os.path.join(os.getcwd(), 'rpdata_scraper'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'rpdata_scraper')
    ]
    
    # Add any additional paths you want to search
    
    for path in search_paths:
        full_path = os.path.join(path, relative_filename)
        if os.path.exists(full_path):
            logging.info(f"File found at: {full_path}")
            return full_path
    
    # If file not found in any location
    logging.error(f"File not found: {relative_filename}")
    return None

def check_zoning_use(all_rows, business_type):
    """
    Determine if properties' zoning allows the specified business type.
    
    Args:
        all_rows (list): List of property data rows from merge_excel.py
        business_type (str): Type of business to check ('Vet' or 'Health')
    
    Returns:
        list: Updated all_rows with 'Allowable Use in Zone (T/F)' filled in
    """
    logger.info("===== CHECKING ZONING ALLOWANCES FOR BUSINESS TYPE: %s =====", business_type)
    
    # Reference file path
    zoning_table_file = get_absolute_file_path("Allowable Use in the Zone - TABLE.xlsx")
    
    if not zoning_table_file:
        logging.error("Could not locate zoning table file")
        return all_rows # Return unchanged rows if file not found
    logger.info(f"Zoning table file located at: {zoning_table_file}")
    
    try:
        # Load the zoning reference table using openpyxl to handle empty cells better
        logger.info(f"Loading zoning reference table from: {zoning_table_file}")
        wb = load_workbook(zoning_table_file)

        print("Workbook Details:")
        print(f"Workbook Title: {zoning_table_file}")
        print(f"Number of Sheets: {len(wb.sheetnames)}")
        print("Sheet Names:", wb.sheetnames)

        ws = wb.active
        
        # Determine which column to use based on business_type
        business_col_idx = None
        if business_type.lower() == 'vet':
            business_col_idx = 2  # Column B (1-indexed)
            logger.info("Using 'Vet' column (B) for zoning check")
        elif business_type.lower() == 'health':
            business_col_idx = 3  # Column C (1-indexed)
            logger.info("Using 'Health' column (C) for zoning check")
        else:
            logger.warning(f"Unknown business type: {business_type}. Cannot determine zoning allowance.")
            return all_rows
        
        # Create a mapping of exact zone descriptions to allowed status
        zone_detail_map = {}
        
        # Start from row 2 (skip header)
        row_idx = 2
        
        # Track consecutive empty rows
        consecutive_empty_rows = 0
        max_empty_rows_before_end = 5  # Allow up to 5 consecutive empty rows
        
        while True:
            # Read zone description from column A
            zone_cell = ws.cell(row=row_idx, column=1)
            zone_value = zone_cell.value
            
            # If we reach an empty row, increment counter but continue
            if zone_value is None or str(zone_value).strip() == "":
                consecutive_empty_rows += 1
                
                # If we've seen too many consecutive empty rows, assume we're at the end
                if consecutive_empty_rows >= max_empty_rows_before_end:
                    logger.info(f"Reached {consecutive_empty_rows} consecutive empty rows, ending processing")
                    break
                    
                # Otherwise, continue to the next row
                row_idx += 1
                continue
            else:
                # Reset consecutive empty rows counter when we find a non-empty row
                consecutive_empty_rows = 0
            
            # Skip rows that are references, dividers, or headers
            zone_str = str(zone_value).strip()
            if (zone_str.startswith("References") or 
                zone_str.startswith("NOTE") or 
                (zone_str.isupper() and len(zone_str.split()) > 1) or  # All caps but not a zone code
                zone_str.startswith("http")):
                row_idx += 1
                continue
            
            # Read allowance value from the business type column
            allowed_cell = ws.cell(row=row_idx, column=business_col_idx)
            
            # Only add to mapping if the cell has a value
            if allowed_cell.value is not None:
                allowed_value = allowed_cell.value
                
                # Convert True/False/string values to consistent format
                if isinstance(allowed_value, bool):
                    formatted_allowed = 'T' if allowed_value else 'F'
                elif isinstance(allowed_value, str):
                    if allowed_value.upper() in ['TRUE', 'YES', 'Y', 'T']:
                        formatted_allowed = 'T'
                    elif allowed_value.upper() in ['FALSE', 'NO', 'N', 'F']:
                        formatted_allowed = 'F'
                    else:
                        # Skip if not a recognized boolean value
                        row_idx += 1
                        continue
                else:
                    # Skip non-boolean values
                    row_idx += 1
                    continue
                
                # Store the full zone string for exact matching
                zone_detail_map[zone_str] = formatted_allowed
                logger.info(f"Added zone: '{zone_str}' with allowed value: '{formatted_allowed}'")
            
            # Move to next row
            row_idx += 1
        
        logger.info(f"Loaded {len(zone_detail_map)} zone descriptions with allowance status")
        
        # Debug: Print the zone map
        for zone, allowed in zone_detail_map.items():
            logger.info(f"Zone: {zone} -> Allowed: {allowed}")
        
        # Process each property row
        updated_count = 0
        for i, row in enumerate(all_rows):
            # Ensure row has enough elements
            while len(row) <= 51:
                row.append("")
                
            # Get the Site Zoning value (index 6)
            site_zoning = row[6]
            
            if not site_zoning or site_zoning == "N/A" or site_zoning == "":
                continue
            
            # Clean and normalize the site zoning string
            site_zoning_str = str(site_zoning).strip()
            
            # ONLY do an exact string match - nothing else
            if site_zoning_str in zone_detail_map:
                row[51] = zone_detail_map[site_zoning_str]
                logger.info(f"Exact match found for '{site_zoning_str}': {row[51]}")
                updated_count += 1
            else:
                logger.info(f"No exact match found for: '{site_zoning_str}', leaving empty")
                # Ensure the field is empty
                row[51] = ""
        
        logger.info(f"Updated 'Allowable Use in Zone' for {updated_count} out of {len(all_rows)} properties")
        return all_rows
    
    except Exception as e:
        logger.error(f"Error checking zoning use: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return all_rows




# Test function for standalone testing
if __name__ == "__main__":
    # Create sample data for testing
    test_rows = [
        ["", "", "", "", "", "", "E1 - Local Centre - North Sydney Local Environmental Plan 2013 Map Amendment No 1", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],


    ]
    '''
    T T
    F T
    '' ''
    T ''
    F ''

    
    '''

    # Test with 'Vet' business type
    print("Testing with business_type='Vet':")
    updated_rows = check_zoning_use(test_rows, "Vet")
    print(updated_rows)
    for row in updated_rows:
        print(row[51])
    print("XXX")
    
    # Test with 'Health' business type
    print("\nTesting with business_type='Health':")
    updated_rows = check_zoning_use(test_rows, "Health")
    print(updated_rows)
    for row in updated_rows:
        print(row[51])
    print("XXX")
